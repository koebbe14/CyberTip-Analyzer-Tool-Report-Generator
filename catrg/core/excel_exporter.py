"""Excel export for IP data and evidence summaries.

Uses the shared parser so there is no duplicated extraction logic.
"""

from __future__ import annotations

from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Font, Alignment

from catrg.core.parser import (
    ParsedCyberTip,
    IpOccurrence,
)
from catrg.core.ip_lookup import IpLookupService
from catrg.utils.date_utils import format_datetime
from catrg.utils.logger import get_logger

log = get_logger(__name__)


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def export_ip_data(
    filename: str,
    all_ip_data: Dict[str, List[IpOccurrence]],
    ip_service: IpLookupService,
    queried_ips: Optional[set] = None,
) -> None:
    """Export IP analysis to an Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IP Address Analysis"

    headers = [
        "IP Address", "Date/Time", "Port", "IP Event",
        "MaxMind Country", "MaxMind City", "Organization", "Registry",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    queried = queried_ips or set(all_ip_data.keys())
    row = 2
    for ip, occurrences in all_ip_data.items():
        for occ in occurrences:
            ws.cell(row=row, column=1).value = ip
            ws.cell(row=row, column=2).value = format_datetime(occ.datetime) or occ.datetime or "N/A"
            ws.cell(row=row, column=3).value = occ.port or "N/A"
            ws.cell(row=row, column=4).value = occ.event or "N/A"

            if ip in queried:
                result = ip_service.get_cached(ip) or ip_service.lookup(ip)
                if result.geo.error:
                    ws.cell(row=row, column=5).value = f"Error: {result.geo.error}"
                    ws.cell(row=row, column=6).value = "N/A"
                else:
                    ws.cell(row=row, column=5).value = result.geo.country
                    ws.cell(row=row, column=6).value = result.geo.city
                if result.whois.error:
                    ws.cell(row=row, column=7).value = f"Error: {result.whois.error}"
                else:
                    ws.cell(row=row, column=7).value = result.whois.organization
                ws.cell(row=row, column=8).value = result.whois.registry
            else:
                for c in (5, 6, 7, 8):
                    ws.cell(row=row, column=c).value = "Not queried (IP limit applied)"
            row += 1

    _auto_width(ws)
    wb.save(filename)
    log.info("IP data exported to %s", filename)


def export_evidence(
    filename: str,
    tip: ParsedCyberTip,
    tips: Optional[List[ParsedCyberTip]] = None,
) -> None:
    """Export evidence summary to an Excel workbook.

    When *tips* is provided (multi-tip mode), a CyberTip ID column is
    prepended and all tips are included.  Otherwise falls back to the
    single-*tip* path for backward compatibility.
    """
    all_tips = tips if tips else [tip]
    multi = len(all_tips) > 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evidence Summary"

    headers = []
    if multi:
        headers.append("CyberTip ID")
    headers.extend([
        "File Number", "File Name", "NCMEC Identifier", "MD5 Hash",
        "Sent Date", "Emailed From", "Emailed To", "Original Filename",
        "Additional Information", "Viewed by ESP", "Upload Date/Time",
        "NCMEC Tags", "Upload IP Address",
        "Webpage Number", "URL", "Type", "Text",
    ])
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    off = 1 if multi else 0
    row = 2

    for t in all_tips:
        esp = t.esp_name

        if ("X Corp" in esp or "Reddit" in esp) and t.webpages:
            for wp in t.webpages:
                if multi:
                    ws.cell(row=row, column=1).value = t.report_id
                ws.cell(row=row, column=14 + off).value = f"Webpage {wp.number}"
                ws.cell(row=row, column=15 + off).value = wp.url or "N/A"
                ws.cell(row=row, column=16 + off).value = wp.type_info or "N/A"
                ws.cell(row=row, column=17 + off).value = wp.text or "N/A"
                if "Reddit" in esp and wp.additional_info:
                    ws.cell(row=row, column=9 + off).value = wp.additional_info
                row += 1

        for ef in t.evidence_files:
            if multi:
                ws.cell(row=row, column=1).value = t.report_id
            ws.cell(row=row, column=1 + off).value = ef.file_number
            ws.cell(row=row, column=2 + off).value = ef.filename or "N/A"
            ws.cell(row=row, column=3 + off).value = ef.submittal_id or "N/A"
            ws.cell(row=row, column=4 + off).value = ef.verification_hash or "N/A"
            ws.cell(row=row, column=5 + off).value = ef.sent_date or "N/A"
            ws.cell(row=row, column=6 + off).value = ef.from_email or "N/A"
            ws.cell(row=row, column=7 + off).value = ef.to_email or "N/A"
            ws.cell(row=row, column=8 + off).value = ef.original_filename or "N/A"
            ws.cell(row=row, column=9 + off).value = ef.additional_info or "N/A"

            if ef.viewed_by_esp is True:
                ws.cell(row=row, column=10 + off).value = "Yes"
            elif ef.viewed_by_esp is False:
                ws.cell(row=row, column=10 + off).value = "No"
            else:
                ws.cell(row=row, column=10 + off).value = "Unknown"

            ws.cell(row=row, column=11 + off).value = format_datetime(ef.upload_time) or ef.upload_time or "N/A"
            ws.cell(row=row, column=12 + off).value = ef.ncmec_tags or "N/A"
            ws.cell(row=row, column=13 + off).value = "; ".join(ef.ip_addresses) if ef.ip_addresses else "N/A"

            row += 1

    _auto_width(ws)
    wb.save(filename)
    log.info("Evidence data exported to %s", filename)
